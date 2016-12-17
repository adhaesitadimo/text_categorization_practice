from collections import defaultdict
from stemming.porter2 import stem
from stop_words import get_stop_words
import codecs
import math
import openpyxl
import os
import re


def tf_builder(text_to_read):
    term_freq = defaultdict(int)
    text_read = text_to_read.read().lower()
    text_read = re.sub('\W', ' ', text_read)
    sum = 0
    for word in text_read.split():
        if (stem(word)) not in stop_words:
            term_freq[stem(word)] += 1
            sum += 1
    sorted_term_freq = list(term_freq.items())
    sorted_term_freq.sort(key=lambda item: item[1], reverse=True)
    return sorted_term_freq


def wf_builder(tf):
    wf = []
    for item in tf:
        cur = list(item)
        cur[1] = math.log(cur[1]) + 1
        wf.append(cur)
    return wf


def idf_builder(tf, docs_amount):
    document_frequencies = defaultdict(int)
    for i in tf:
        for j in i:
            document_frequencies[j[0]] += 1
    for key in document_frequencies:
        document_frequencies[key] = math.log(docs_amount/document_frequencies[key])
    return document_frequencies


def tfidf_builder(tf, idf):
    tfidf = defaultdict(int)
    for item_tf in tf:
        for key in idf:
            if item_tf[0] == key:
                tfidf[item_tf[0]] = item_tf[1] * idf[key]
    sorted_tfidf = list(tfidf.items())
    sorted_tfidf.sort(key=lambda item: item[1], reverse=True)
    return sorted_tfidf


wb = openpyxl.load_workbook(filename='C:/table/tfidf.xlsx')
directory = 'C:/texts/'
cur_text = ''
stop_words = get_stop_words('en')
l = os.listdir(directory)
filenames = []
filenames_corrected = []
term_frequencies = []
weight_frequencies = []
tfidfs = []
wfidfs = []
docs_amount = 0
for i in l:
    file = codecs.open(i, 'r', 'cp1251')
    filenames.append(i)
    docs_amount += 1
    term_freq = tf_builder(file)
    weight_freq = wf_builder(term_freq)
    term_frequencies.append(term_freq)
    weight_frequencies.append(weight_freq)
idf = idf_builder(term_frequencies, docs_amount)
for i in term_frequencies:
    current_tfidf = tfidf_builder(i, idf)
    tfidfs.append(current_tfidf)
for i in weight_frequencies:
    current_wfidf = tfidf_builder(i, idf)
    wfidfs.append(current_wfidf)
for item in filenames:
    filenames_corrected.append(item.replace('.txt', ''))
terms = idf.keys()
terms = list(terms)
terms.sort()
sheet = wb['TF-IDF']
row_num = 2
for rec in terms:
    sheet.cell(row=row_num, column=1).value = rec
    row_num += 1
column_num = 2
for rec in filenames_corrected:
    sheet.cell(row=1, column=column_num).value = rec
    column_num += 1
row_num = 1
column_num = 1
for rec_row in terms:
    row_num += 1
    column_num = 1
    for rec_column in tfidfs:
        column_num += 1
        sheet.cell(row=row_num, column=column_num).value = 0
row_num = 1
column_num = 1
for item in terms:
    row_num += 1
    column_num = 1
    for tfidf in tfidfs:
        column_num += 1
        for j in tfidf:
            if item == j[0]:
                print(item)
                print(row_num, column_num, j[1])
                sheet.cell(row=row_num, column=column_num).value = j[1]
sheet = wb['WF-IDF']
row_num = 2
for rec in terms:
    sheet.cell(row=row_num, column=1).value = rec
    row_num += 1
column_num = 2
for rec in filenames_corrected:
    sheet.cell(row=1, column=column_num).value = rec
    column_num += 1
row_num = 1
column_num = 1
for i in terms:
    row_num += 1
    column_num = 1
    for j in wfidfs:
        column_num += 1
        sheet.cell(row=row_num, column=column_num).value = 0
row_num = 1
column_num = 1
for item in terms:
    row_num += 1
    column_num = 1
    for wfidf in wfidfs:
        column_num += 1
        for j in wfidf:
            if item == j[0]:
                print(item)
                print(row_num, column_num, j[1])
                sheet.cell(row=row_num, column=column_num).value = j[1]
wb.save('C:/table/tfidf.xlsx')
